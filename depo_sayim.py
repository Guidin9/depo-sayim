#!/usr/bin/env python3
"""
Depo Sayim Uygulamasi - Faz 1
Kullanim:
    python depo_sayim.py yukle <rapor.xlsx>   # Tiger Seri/Lot Envanter Raporu'nu yukle
    python depo_sayim.py                       # sunucuyu baslat -> http://localhost:8000
Gereksinim: Python 3.8+, openpyxl  (pip install openpyxl)
"""
import http.server, socketserver, sqlite3, json, re, sys, os, urllib.parse, datetime

DB = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sayim.db")
PORT = 8000

# ---------------------------------------------------------------- normalizasyon
TR = str.maketrans("ıİğĞüÜşŞöÖçÇ", "iigguussoocc")

def norm(s):
    """Buyuk harf, Turkce katla, harf-rakam disini at."""
    if s is None:
        return ""
    return re.sub(r"[^A-Z0-9]", "", str(s).translate(TR).upper())

KOMUT = {"##SONRAKI##": "sonraki", "##IPTAL##": "iptal", "##GERIAL##": "gerial",
         "##FAZLA##": "fazla", "##ATLA##": "atla", "##BITIR##": "bitir"}

def upc_mi(s):
    """12-13 haneli, kontrol hanesi tutan perakende barkodu mu?"""
    s = re.sub(r"\D", "", str(s))
    if len(s) not in (12, 13):
        return False
    d = [int(c) for c in s]
    t = sum(d[i] * (3 if (len(s) - 2 - i) % 2 == 0 else 1) for i in range(len(s) - 1))
    return (10 - t % 10) % 10 == d[-1]

KIRLI_KELIME = re.compile(r"SAYIM|SAYIN|STOK|CIKAN|DENEME|FAZLA|TEST|PROJE|DEPO|BAKIM")

def kirli_mi(seri, kod):
    """Seri numarasi gercek mi, yoksa stok tutturmak icin uydurulmus mu?"""
    n, k = norm(seri), norm(kod)
    if not n:
        return 1, "bos"
    if re.search(r"[ ]", str(seri)):
        return 1, "bosluk"
    if KIRLI_KELIME.search(n):
        return 1, "placeholder"
    if k and len(k) > 3 and n.startswith(k):
        return 1, "kod+sayac"
    if len(n) > 25:
        return 1, "asiri uzun"
    return 0, ""

# ---------------------------------------------------------------- veritabani
SEMA = """
CREATE TABLE IF NOT EXISTS beklenen(
  id INTEGER PRIMARY KEY, kod TEXT, aciklama TEXT, tur TEXT, ambar TEXT,
  izleme TEXT, seri TEXT, seri_n TEXT, miktar REAL, birim TEXT,
  kirli INT, kirli_sebep TEXT);
CREATE INDEX IF NOT EXISTS ix_seri ON beklenen(seri_n);
CREATE INDEX IF NOT EXISTS ix_kod  ON beklenen(kod);

CREATE TABLE IF NOT EXISTS oturum(
  id INTEGER PRIMARY KEY, ambar TEXT, basla TEXT, bitir TEXT);

CREATE TABLE IF NOT EXISTS okutma(
  id INTEGER PRIMARY KEY, oturum INT, ts TEXT, ham TEXT,
  kod TEXT, seri TEXT, miktar REAL, beklenen_id INT, tip TEXT, not_ TEXT);

CREATE TABLE IF NOT EXISTS eslesme(
  barkod TEXT PRIMARY KEY, kod TEXT, seri TEXT, ts TEXT);

CREATE TABLE IF NOT EXISTS tampon(
  id INTEGER PRIMARY KEY, oturum INT, ts TEXT, ham TEXT);

CREATE TABLE IF NOT EXISTS kuyruk(
  id INTEGER PRIMARY KEY, oturum INT, ts TEXT, barkodlar TEXT,
  raf TEXT, cozuldu INT DEFAULT 0);
"""

def db():
    c = sqlite3.connect(DB)
    c.row_factory = sqlite3.Row
    c.executescript(SEMA)
    return c

# ---------------------------------------------------------------- excel yukleme
def yukle(yol):
    from openpyxl import load_workbook
    wb = load_workbook(yol, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    basliklar = None
    for r in rows:
        if r and "Malzeme Kodu" in [str(x).strip() if x else "" for x in r]:
            basliklar = [str(x).strip() if x else "" for x in r]
            break
    if not basliklar:
        sys.exit("HATA: 'Malzeme Kodu' basligi bulunamadi. Dogru rapor mu?")
    ix = {h: i for i, h in enumerate(basliklar)}

    def g(r, ad):
        i = ix.get(ad)
        return r[i] if i is not None and i < len(r) else None

    c = db()
    c.execute("DELETE FROM beklenen")
    n = 0
    for r in rows:
        kod = g(r, "Malzeme Kodu")
        if not kod:
            continue
        seri = g(r, "Seri/Lot No.")
        izl = str(g(r, "İzleme Yöntemi") or "")
        izleme = "seri" if "Seri" in izl else ("lot" if "Lot" in izl else "yok")
        k, sb = kirli_mi(seri, kod) if izleme == "seri" else (0, "")
        c.execute("""INSERT INTO beklenen(kod,aciklama,tur,ambar,izleme,seri,seri_n,
                     miktar,birim,kirli,kirli_sebep) VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                  (str(kod).strip(), str(g(r, "Malzeme Açıklaması") or ""),
                   str(g(r, "Malzeme Türü") or ""),
                   str(g(r, "Ambar Maliyet Grubu") or ""), izleme,
                   str(seri or ""), norm(seri), float(g(r, "Envanter Miktarı") or 0),
                   str(g(r, "Birim") or "AD"), k, sb))
        n += 1
    c.commit()

    s = c.execute("""SELECT izleme, COUNT(*) a, SUM(miktar) m, SUM(kirli) k
                     FROM beklenen GROUP BY izleme""").fetchall()
    print(f"Yuklendi: {n} satir\n")
    for r in s:
        print(f"  {r['izleme']:5} satir={r['a']:4}  adet={int(r['m'] or 0):5}  kirli={r['k'] or 0}")
    amb = c.execute("SELECT ambar, COUNT(*) a FROM beklenen GROUP BY ambar").fetchall()
    print("\n  Ambarlar: " + ", ".join(f"{r['ambar']}({r['a']})" for r in amb))
    c.close()

# ---------------------------------------------------------------- eslestirme motoru
def eslestir(c, ham, ambar, oturum):
    """Okutulan kodu beklenen listeyle eslestirir. Sonuc dict doner."""
    n = norm(ham)
    if not n:
        return {"tip": "bos"}

    sayilan = set(r["beklenen_id"] for r in c.execute(
        "SELECT beklenen_id FROM okutma WHERE oturum=? AND beklenen_id IS NOT NULL", (oturum,)))

    # 1) Birebir seri eslesmesi
    for r in c.execute("SELECT * FROM beklenen WHERE seri_n=? AND ambar=?", (n, ambar)):
        if r["id"] in sayilan:
            return {"tip": "tekrar", "kod": r["kod"], "aciklama": r["aciklama"],
                    "seri": r["seri"]}
        return {"tip": "eslesti", "id": r["id"], "kod": r["kod"],
                "aciklama": r["aciklama"], "seri": r["seri"]}

    # 2) Ogrenilmis eslesme
    e = c.execute("SELECT * FROM eslesme WHERE barkod=?", (n,)).fetchone()
    if e:
        r = c.execute("SELECT * FROM beklenen WHERE kod=? AND ambar=? AND id NOT IN "
                      "(SELECT COALESCE(beklenen_id,-1) FROM okutma WHERE oturum=?)",
                      (e["kod"], ambar, oturum)).fetchone()
        if r:
            return {"tip": "eslesti", "id": r["id"], "kod": r["kod"],
                    "aciklama": r["aciklama"], "seri": r["seri"], "not": "ogrenilmis"}

    # 3) Icerme: gercek S/N kirli kaydin icine gomulmus olabilir
    if len(n) >= 6:
        adaylar = []
        for r in c.execute("SELECT * FROM beklenen WHERE ambar=? AND izleme='seri'", (ambar,)):
            if r["id"] in sayilan:
                continue
            if n in r["seri_n"] and norm(r["kod"]) != n:
                adaylar.append(dict(id=r["id"], kod=r["kod"], aciklama=r["aciklama"],
                                    seri=r["seri"]))
        if adaylar:
            return {"tip": "aday", "adaylar": adaylar[:8]}

    # 4) Malzeme kodu eslesmesi -> miktar modu / bos slot
    kodlar = [r for r in c.execute("SELECT * FROM beklenen WHERE ambar=?", (ambar,))
              if norm(r["kod"]) == n]
    if kodlar:
        acik = [dict(id=r["id"], seri=r["seri"], kirli=r["kirli"])
                for r in kodlar if r["id"] not in sayilan]
        return {"tip": "kod", "kod": kodlar[0]["kod"], "aciklama": kodlar[0]["aciklama"],
                "izleme": kodlar[0]["izleme"], "birim": kodlar[0]["birim"],
                "acik": acik, "acik_kirli": sum(1 for a in acik if a["kirli"])}

    # 5) Bilinmiyor
    return {"tip": "bilinmiyor", "ham": ham}

# ---------------------------------------------------------------- tekil cozumleme
def coz(c, ham, ambar, sayilan):
    """Tek bir okutmayi cozer: seri / kod / ogrenilmis / upc / bilinmiyor."""
    n = norm(ham)
    if not n:
        return {"t": "bos"}

    r = c.execute("SELECT * FROM beklenen WHERE seri_n=? AND ambar=?", (n, ambar)).fetchone()
    if r:
        return {"t": "tekrar" if r["id"] in sayilan else "seri", "id": r["id"],
                "kod": r["kod"], "aciklama": r["aciklama"], "seri": r["seri"]}

    for r in c.execute("SELECT * FROM beklenen WHERE ambar=?", (ambar,)):
        if norm(r["kod"]) == n:
            return {"t": "kod", "kod": r["kod"], "aciklama": r["aciklama"],
                    "izleme": r["izleme"], "birim": r["birim"]}

    # kod onek eslesmesi: ARK-1250L-S5A1 -> ARK1250LS5A1ATR8641924
    if len(n) >= 8:
        for r in c.execute("SELECT * FROM beklenen WHERE ambar=?", (ambar,)):
            k = norm(r["kod"])
            if len(k) >= 8 and (k.startswith(n) or n.startswith(k)):
                return {"t": "kod", "kod": r["kod"], "aciklama": r["aciklama"],
                        "izleme": r["izleme"], "birim": r["birim"], "not": "önek eşleşmesi"}

    e = c.execute("SELECT * FROM eslesme WHERE barkod=?", (n,)).fetchone()
    if e:
        r = c.execute("SELECT * FROM beklenen WHERE kod=? AND ambar=?",
                      (e["kod"], ambar)).fetchone()
        if r:
            return {"t": "ogrenilmis", "kod": r["kod"], "aciklama": r["aciklama"],
                    "izleme": r["izleme"], "birim": r["birim"]}

    # icerme: gercek S/N kirli kaydin icine gomulmus
    if len(n) >= 6:
        for r in c.execute("SELECT * FROM beklenen WHERE ambar=? AND izleme='seri' AND kirli=1",
                           (ambar,)):
            if r["id"] not in sayilan and n in r["seri_n"] and norm(r["kod"]) != n:
                return {"t": "seri", "id": r["id"], "kod": r["kod"],
                        "aciklama": r["aciklama"], "seri": r["seri"], "not": "gömülü S/N"}

    return {"t": "upc" if upc_mi(ham) else "bilinmiyor", "ham": ham}


# ---------------------------------------------------------------- grup cozumleme
def grup_coz(c, oturum, ambar, raf=None):
    """Tampondaki barkodlari TEK URUN kabul edip cozer. ##SONRAKI## okutunca calisir."""
    ts = datetime.datetime.now().isoformat()
    hamlar = [r["ham"] for r in c.execute(
        "SELECT ham FROM tampon WHERE oturum=? ORDER BY id", (oturum,))]
    c.execute("DELETE FROM tampon WHERE oturum=?", (oturum,))
    if not hamlar:
        return {"tip": "bos"}

    sayilan = set(r["beklenen_id"] for r in c.execute(
        "SELECT beklenen_id FROM okutma WHERE oturum=? AND beklenen_id IS NOT NULL", (oturum,)))
    coz_list = [(h, coz(c, h, ambar, sayilan)) for h in hamlar]

    seri_h = next((x for x in coz_list if x[1]["t"] == "seri"), None)
    kod_h = next((x for x in coz_list if x[1]["t"] in ("kod", "ogrenilmis")), None)
    tekrar = next((x for x in coz_list if x[1]["t"] == "tekrar"), None)
    bilinmeyen = [h for h, r in coz_list if r["t"] in ("bilinmiyor", "upc")]

    if tekrar and not seri_h:
        return {"tip": "tekrar", "kod": tekrar[1]["kod"], "seri": tekrar[1]["seri"],
                "ses": "uyari"}

    kaynak = seri_h or kod_h
    if not kaynak:
        c.execute("INSERT INTO kuyruk(oturum,ts,barkodlar,raf) VALUES(?,?,?,?)",
                  (oturum, ts, json.dumps(hamlar, ensure_ascii=False), raf))
        return {"tip": "kuyruk", "barkodlar": hamlar, "ses": "kuyruk"}

    kod = kaynak[1]["kod"]
    aciklama = kaynak[1]["aciklama"]

    # bilinmeyen barkodlari bu malzemeye ogret
    ogrenilen = []
    for h in bilinmeyen:
        c.execute("INSERT OR REPLACE INTO eslesme VALUES(?,?,?,?)", (norm(h), kod, "", ts))
        ogrenilen.append(h)

    if seri_h:
        r = seri_h[1]
        c.execute("""INSERT INTO okutma(oturum,ts,ham,kod,seri,miktar,beklenen_id,tip,not_)
                     VALUES(?,?,?,?,?,1,?,'eslesti',?)""",
                  (oturum, ts, seri_h[0], kod, r["seri"], r["id"],
                   (r.get("not") or "") + (f" | öğrenildi: {','.join(ogrenilen)}" if ogrenilen else "")))
        return {"tip": "eslesti", "kod": kod, "aciklama": aciklama, "seri": r["seri"],
                "ogrenilen": ogrenilen, "ses": "ok"}

    # Malzeme belli ama seri eslesmedi
    izleme = kaynak[1].get("izleme", "yok")
    if izleme == "seri":
        slot = c.execute("""SELECT * FROM beklenen WHERE kod=? AND ambar=? AND kirli=1
                            AND id NOT IN (SELECT COALESCE(beklenen_id,-1) FROM okutma
                            WHERE oturum=?) LIMIT 1""", (kod, ambar, oturum)).fetchone()
        yeni_sn = max(bilinmeyen, key=len) if bilinmeyen else ""
        if slot:
            c.execute("""INSERT INTO okutma(oturum,ts,ham,kod,seri,miktar,beklenen_id,tip,not_)
                         VALUES(?,?,?,?,?,1,?,'eslesti','slot dolduruldu')""",
                      (oturum, ts, yeni_sn or kod, kod, slot["seri"], slot["id"]))
            return {"tip": "slot", "kod": kod, "aciklama": aciklama, "eski": slot["seri"],
                    "yeni": yeni_sn, "ses": "ok"}
        c.execute("""INSERT INTO okutma(oturum,ts,ham,kod,seri,miktar,tip,not_)
                     VALUES(?,?,?,?,?,1,'fazla','seri takipli, karşılığı yok')""",
                  (oturum, ts, yeni_sn or kod, kod, yeni_sn))
        return {"tip": "fazla", "kod": kod, "aciklama": aciklama, "ses": "uyari"}

    b = c.execute("SELECT * FROM beklenen WHERE kod=? AND ambar=? LIMIT 1", (kod, ambar)).fetchone()
    c.execute("""INSERT INTO okutma(oturum,ts,ham,kod,seri,miktar,beklenen_id,tip,not_)
                 VALUES(?,?,?,?,?,1,?,'kod','adet +1')""",
              (oturum, ts, kaynak[0], kod, b["seri"] if b else "", b["id"] if b else None))
    top = c.execute("SELECT SUM(miktar) s FROM okutma WHERE oturum=? AND kod=?",
                    (oturum, kod)).fetchone()["s"]
    return {"tip": "adet", "kod": kod, "aciklama": aciklama, "toplam": top,
            "beklenen": b["miktar"] if b else 0, "ogrenilen": ogrenilen, "ses": "ok"}


# ---------------------------------------------------------------- rapor
def rapor(oturum_id, yol):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    c = db()
    o = c.execute("SELECT * FROM oturum WHERE id=?", (oturum_id,)).fetchone()
    if not o:
        sys.exit("Oturum bulunamadi")
    amb = o["ambar"]
    wb = Workbook()
    kalin = Font(name="Arial", bold=True)

    def sayfa(ad, basliklar, satirlar):
        ws = wb.create_sheet(ad)
        ws.append(basliklar)
        for h in ws[1]:
            h.font = kalin
        for s in satirlar:
            ws.append(list(s))
        for i, b in enumerate(basliklar, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = max(12, min(45, len(b) + 18))
        return ws

    sayilan = set(r["beklenen_id"] for r in c.execute(
        "SELECT beklenen_id FROM okutma WHERE oturum=? AND beklenen_id IS NOT NULL", (oturum_id,)))

    eksik = [(r["kod"], r["aciklama"], r["seri"], r["miktar"], r["birim"],
              "KIRLI KAYIT" if r["kirli"] else "")
             for r in c.execute("SELECT * FROM beklenen WHERE ambar=?", (amb,))
             if r["id"] not in sayilan]

    fazla = [(r["ts"][:19], r["ham"], r["kod"] or "?", r["seri"] or "", r["miktar"], r["not_"] or "")
             for r in c.execute("SELECT * FROM okutma WHERE oturum=? AND tip IN ('fazla','bilinmiyor')",
                                (oturum_id,))]

    eslesen = [(r["ts"][:19], r["kod"], r["seri"], r["miktar"], r["tip"])
               for r in c.execute("SELECT * FROM okutma WHERE oturum=? AND tip IN ('eslesti','kod')",
                                  (oturum_id,))]

    duzeltme = []
    for r in c.execute("""SELECT o.ham, o.ts, b.kod, b.aciklama, b.seri
                          FROM okutma o JOIN beklenen b ON b.id=o.beklenen_id
                          WHERE o.oturum=? AND b.kirli=1 AND o.ham<>''""", (oturum_id,)):
        duzeltme.append((r["kod"], r["aciklama"], r["seri"], r["ham"], r["ts"][:19]))

    barkodlar = [(r["barkod"], r["kod"],
                  (c.execute("SELECT aciklama FROM beklenen WHERE kod=? LIMIT 1",
                             (r["kod"],)).fetchone() or {"aciklama": ""})["aciklama"],
                  r["ts"][:19])
                 for r in c.execute("SELECT * FROM eslesme ORDER BY kod")]

    sayfa("Eksik", ["Malzeme Kodu", "Açıklama", "Beklenen Seri/Lot", "Miktar", "Birim", "Not"], eksik)
    sayfa("Fazla", ["Zaman", "Okutulan", "Malzeme Kodu", "Seri", "Miktar", "Not"], fazla)
    sayfa("Eşleşen", ["Zaman", "Malzeme Kodu", "Seri/Lot", "Miktar", "Tip"], eslesen)
    ws = sayfa("Tiger Düzeltme", ["Malzeme Kodu", "Açıklama", "MEVCUT (hatalı) Seri No",
                                 "YENİ (gerçek) Seri No", "Zaman"], duzeltme)
    ws.append([])
    ws.append(["Bu sayfadaki kayıtlar Tiger'da seri numarası düzeltmesi gerektirir.",
               "Ambar Sayımı ekranından fiş oluştururken kullanın."])
    ws = sayfa("Barkod Tablosu", ["Okutulan Barkod", "Malzeme Kodu", "Açıklama", "Öğrenildiği An"],
               barkodlar)
    ws.append([])
    ws.append(["Bu barkodları Tiger'da malzeme kartı > Birimler > Barkod alanına yazın.",
               "Yazdıktan sonra bu ürünler sorusuz eşleşir."])
    del wb["Sheet"]
    wb.save(yol)
    print(f"Rapor yazildi: {yol}")
    print(f"  Eksik={len(eksik)}  Fazla={len(fazla)}  Eslesen={len(eslesen)}  "
          f"Duzeltme={len(duzeltme)}  Barkod={len(barkodlar)}")
    c.close()

# ---------------------------------------------------------------- web arayuz
SAYFA = r"""<!doctype html><html lang=tr><meta charset=utf-8>
<meta name=viewport content="width=device-width,initial-scale=1">
<title>Depo Sayim</title><style>
*{box-sizing:border-box}
body{font:15px/1.45 -apple-system,Segoe UI,Roboto,sans-serif;margin:0;background:#12151a;color:#e6e9ef}
header{background:#1a1f28;padding:10px 16px;display:flex;gap:14px;align-items:center;
 border-bottom:1px solid #2a3140;flex-wrap:wrap}
header b{font-size:16px}
.sayac{margin-left:auto;display:flex;gap:16px;font-variant-numeric:tabular-nums}
.sayac div{text-align:center}.sayac span{display:block;font-size:20px;font-weight:600}
.sayac small{color:#8c95a6;font-size:11px;text-transform:uppercase;letter-spacing:.4px}
main{padding:16px;max-width:900px;margin:auto}
#giris{width:100%;padding:16px;font-size:20px;background:#0d1015;border:2px solid #3d5afe;
 border-radius:8px;color:#fff;font-family:ui-monospace,monospace}
#giris:focus{outline:none;border-color:#7c8cff}
.satir{padding:10px 12px;border-radius:6px;margin-top:8px;background:#1a1f28;
 border-left:4px solid #444;display:flex;gap:12px;align-items:baseline}
.ok{border-left-color:#2ecc71}.err{border-left-color:#e74c3c}
.warn{border-left-color:#f39c12}.info{border-left-color:#3498db}
.satir code{font-family:ui-monospace,monospace;color:#7c8cff}
.satir .ac{color:#8c95a6;font-size:13px;flex:1}
button{padding:9px 14px;font-size:14px;border:0;border-radius:6px;background:#3d5afe;
 color:#fff;cursor:pointer}button:hover{background:#5670ff}
button.sec{background:#2a3140}button.sec:hover{background:#38414f}
#kutu{position:fixed;inset:0;background:#000a;display:none;align-items:center;justify-content:center;padding:16px}
#kutu>div{background:#1a1f28;padding:20px;border-radius:10px;max-width:640px;width:100%;
 max-height:82vh;overflow:auto}
#kutu h3{margin:0 0 12px}
#ara{width:100%;padding:11px;font-size:16px;background:#0d1015;border:1px solid #3a4252;
 border-radius:6px;color:#fff;margin-bottom:10px}
.sec-satir{padding:10px;border-radius:6px;cursor:pointer;border:1px solid #2a3140;margin-bottom:6px}
.sec-satir:hover{background:#2a3140}
.sec-satir b{color:#7c8cff;font-family:ui-monospace,monospace}
.kirli{color:#f39c12;font-size:12px}
</style>
<header>
  <b>Depo Sayım</b>
  <span id=amb style="color:#8c95a6"></span>
  <span id=raf style="color:#f39c12"></span>
  <button class=sec onclick="kuyrukAc()">Kuyruk</button>
  <button class=sec onclick="bitir()">Raporu indir</button>
  <div class=sayac>
    <div><span id=s_ok>0</span><small>okutuldu</small></div>
    <div><span id=s_kalan>0</span><small>kalan</small></div>
    <div><span id=s_fazla>0</span><small>fazla</small></div>
  </div>
</header>
<main>
  <input id=giris autofocus autocomplete=off placeholder="Barkodu okut — ürün bitince SIRADAKİ ÜRÜN barkodunu okut">
  <div id=tampon style="color:#f39c12;font-size:13px;min-height:20px;margin-top:6px"></div>
  <div id=akis></div>
</main>
<div id=kutu><div>
  <h3 id=k_baslik></h3>
  <input id=ara placeholder="kod veya açıklama ara...">
  <div id=k_liste></div>
  <div style="margin-top:12px;display:flex;gap:8px">
    <button class=sec onclick="kapat()">Vazgeç (Esc)</button>
    <button class=sec onclick="fazlaYaz()">Fazla olarak kaydet</button>
  </div>
</div></div>
<script>
let OTURUM=null,AMBAR=null,RAF=null,sonHam=null;
const $=s=>document.querySelector(s);
const g=$('#giris'),akis=$('#akis'),kutu=$('#kutu');

/* --- ses: ekrana bakmadan geri bildirim --- */
let AC=null;
function bip(tip){
  try{ AC=AC||new (window.AudioContext||window.webkitAudioContext)();
    const desen={tik:[[900,.035]],ok:[[1250,.09]],
      uyari:[[500,.12],[0,.05],[500,.12]],
      kuyruk:[[750,.07],[0,.04],[750,.07],[0,.04],[750,.07]],
      bitti:[[900,.1],[0,.05],[1200,.1],[0,.05],[1500,.18]]}[tip]||[[900,.04]];
    let t=AC.currentTime;
    for(const [f,d] of desen){
      if(f){const o=AC.createOscillator(),v=AC.createGain();
        o.frequency.value=f;o.type='square';v.gain.value=.14;
        o.connect(v);v.connect(AC.destination);o.start(t);o.stop(t+d);}
      t+=d;}
  }catch(e){}
}

async function api(u,d){const r=await fetch(u,{method:'POST',
  headers:{'Content-Type':'application/json'},body:JSON.stringify(d)});return r.json()}

function yaz(sinif,ana,alt){const d=document.createElement('div');
  d.className='satir '+sinif;
  d.innerHTML=`<code>${ana}</code><span class=ac>${alt||''}</span>`;
  akis.prepend(d);while(akis.children.length>40)akis.lastChild.remove()}

async function sayac(){const s=await api('/api/sayac',{oturum:OTURUM});
  $('#s_ok').textContent=s.ok;$('#s_kalan').textContent=s.kalan;$('#s_fazla').textContent=s.fazla}

g.addEventListener('keydown',async e=>{
  if(e.key!=='Enter')return;
  const v=g.value.trim();g.value='';if(!v)return;
  sonHam=v;
  const r=await api('/api/okut',{oturum:OTURUM,ham:v,raf:RAF});
  bip(r.ses||'tik');
  isle(r,v);
});

function isle(r,v){
  const t=r.tip;
  if(t==='tampon'){
    const et={seri:'S/N tanındı',kod:'malzeme kodu',ogrenilmis:'öğrenilmiş barkod',
      upc:'UPC barkodu',bilinmiyor:'tanınmadı',tekrar:'bu S/N zaten okutuldu'}[r.coz]||r.coz;
    $('#tampon').textContent=`grupta ${r.adet} barkod`;
    yaz('info','· '+v,et+(r.kod?' → '+r.kod:''));
  }
  else if(t==='eslesti'){$('#tampon').textContent='';
    yaz('ok','✓ '+r.kod,r.aciklama+' — '+r.seri+
      (r.ogrenilen&&r.ogrenilen.length?' | öğrenildi: '+r.ogrenilen.join(', '):''))}
  else if(t==='slot'){$('#tampon').textContent='';
    yaz('ok','✓ '+r.kod,'uydurma kayıt düzeltildi: '+r.eski+' → '+(r.yeni||'?'))}
  else if(t==='adet'){$('#tampon').textContent='';
    yaz('ok','✓ '+r.kod,r.aciklama+' — sayılan '+r.toplam+' / beklenen '+r.beklenen)}
  else if(t==='fazla'||t==='fazla_elle'){$('#tampon').textContent='';
    yaz('err','⚠ FAZLA',(r.kod||'')+' '+(r.barkodlar||[]).join(', '))}
  else if(t==='tekrar'){$('#tampon').textContent='';
    yaz('warn','⟳ tekrar',r.kod+' — '+r.seri+' zaten okutuldu')}
  else if(t==='kuyruk'){$('#tampon').textContent='';
    yaz('warn','? kuyruğa atıldı',(r.barkodlar||[]).join(' + ')+' — sayım sonunda çözülecek')}
  else if(t==='iptal'){$('#tampon').textContent='';yaz('warn','↺ grup iptal','')}
  else if(t==='gerial'){yaz('warn','← silindi',r.ham||'')}
  else if(t==='raf'){RAF=r.raf;$('#raf').textContent='Raf '+RAF;yaz('info','▣ raf '+RAF,'')}
  else if(t==='bitti'){yaz('ok','■ sayım bitti','rapor hazırlanıyor...');
    setTimeout(()=>location.href='/rapor.xlsx?oturum='+OTURUM,600)}
  sayac();
}

/* --- kuyruk cozumleme --- */
let aktifK=null;
async function kuyrukAc(){
  const r=await api('/api/kuyruk',{oturum:OTURUM});
  if(!r.kuyruk.length){alert('Kuyruk boş.');return}
  ac('Çözülmeyi bekleyen '+r.kuyruk.length+' ürün');
  $('#k_liste').innerHTML=r.kuyruk.map(q=>
    `<div class=sec-satir onclick='kSec(${q.id},${JSON.stringify(JSON.stringify(q.barkodlar))})'>
      <b>${q.barkodlar.join(' + ')}</b><br>
      <span class=kirli>${q.raf?'raf '+q.raf+' · ':''}${q.ts}</span></div>`).join('');
}
function kSec(id,bs){aktifK=id;
  $('#k_baslik').textContent=JSON.parse(bs).join(' + ')+' — hangi malzeme?';
  $('#k_liste').innerHTML='<p style="color:#8c95a6">Aramaya başlayın...</p>';
  $('#ara').value='';$('#ara').focus()}

$('#ara').addEventListener('input',async e=>{
  const q=e.target.value.trim();if(q.length<2)return;
  const r=await api('/api/ara',{ambar:AMBAR,q:q});
  $('#k_liste').innerHTML=r.sonuc.map(a=>
    `<div class=sec-satir onclick='kBagla(${a.id})'>
      <b>${a.kod}</b> — ${a.aciklama}<br>
      <span class=kirli>${a.seri||''} ${a.kirli?'(uydurma kayıt)':''}</span></div>`).join('')
    ||'<p style="color:#8c95a6">Sonuç yok</p>'});

async function kBagla(bid){
  await api('/api/kuyruk_coz',{id:aktifK,bid:bid});
  bip('ok');kuyrukAc();sayac()}

function ac(b){$('#k_baslik').textContent=b;kutu.style.display='flex';$('#ara').value=''}
function kapat(){kutu.style.display='none';g.focus()}
document.addEventListener('keydown',e=>{if(e.key==='Escape')kapat()});
document.addEventListener('click',e=>{if(e.target===kutu)kapat()});
async function bitir(){location.href='/rapor.xlsx?oturum='+OTURUM}

(async()=>{const r=await api('/api/basla',{});
  OTURUM=r.oturum;AMBAR=r.ambar;
  $('#amb').textContent='Ambar '+AMBAR+' · oturum #'+OTURUM;
  sayac();g.focus()})();
</script></html>"""

class H(http.server.BaseHTTPRequestHandler):
    def log_message(self, *a):
        pass

    def _j(self, o, kod=200):
        b = json.dumps(o, ensure_ascii=False).encode()
        self.send_response(kod)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(b)))
        self.end_headers()
        self.wfile.write(b)

    def do_GET(self):
        if self.path.startswith("/rapor.xlsx"):
            q = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
            oid = int(q.get("oturum", ["0"])[0])
            yol = os.path.join(os.path.dirname(DB), f"sayim_raporu_{oid}.xlsx")
            rapor(oid, yol)
            with open(yol, "rb") as f:
                b = f.read()
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f'attachment; filename="sayim_raporu_{oid}.xlsx"')
            self.send_header("Content-Length", str(len(b)))
            self.end_headers()
            self.wfile.write(b)
            return
        b = SAYFA.encode()
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(b)))
        self.end_headers()
        self.wfile.write(b)

    def do_POST(self):
        n = int(self.headers.get("Content-Length", 0))
        d = json.loads(self.rfile.read(n) or b"{}")
        c = db()
        try:
            self._j(self.yonlendir(c, d))
        finally:
            c.commit()
            c.close()

    def yonlendir(self, c, d):
        p = self.path
        ts = datetime.datetime.now().isoformat()

        if p == "/api/basla":
            o = c.execute("SELECT * FROM oturum WHERE bitir IS NULL ORDER BY id DESC").fetchone()
            if o:
                return {"oturum": o["id"], "ambar": o["ambar"]}
            amb = c.execute("SELECT ambar FROM beklenen GROUP BY ambar "
                            "ORDER BY COUNT(*) DESC").fetchone()
            amb = amb["ambar"] if amb else "1"
            cur = c.execute("INSERT INTO oturum(ambar,basla) VALUES(?,?)", (amb, ts))
            return {"oturum": cur.lastrowid, "ambar": amb}

        if p == "/api/sayac":
            o = d["oturum"]
            amb = c.execute("SELECT ambar FROM oturum WHERE id=?", (o,)).fetchone()["ambar"]
            top = c.execute("SELECT COUNT(*) n FROM beklenen WHERE ambar=?", (amb,)).fetchone()["n"]
            ok = c.execute("SELECT COUNT(DISTINCT beklenen_id) n FROM okutma "
                           "WHERE oturum=? AND beklenen_id IS NOT NULL", (o,)).fetchone()["n"]
            fz = c.execute("SELECT COUNT(*) n FROM okutma WHERE oturum=? AND tip IN "
                           "('fazla','bilinmiyor')", (o,)).fetchone()["n"]
            return {"ok": ok, "kalan": top - ok, "fazla": fz}

        if p == "/api/okut":
            o, ham = d["oturum"], d["ham"].strip()
            row = c.execute("SELECT * FROM oturum WHERE id=?", (o,)).fetchone()
            amb = row["ambar"]
            U = ham.upper()

            if U in KOMUT:
                k = KOMUT[U]
                if k == "sonraki":
                    return grup_coz(c, o, amb, d.get("raf"))
                if k == "iptal":
                    c.execute("DELETE FROM tampon WHERE oturum=?", (o,))
                    return {"tip": "iptal", "ses": "uyari"}
                if k == "gerial":
                    l = c.execute("SELECT id,ham FROM tampon WHERE oturum=? ORDER BY id DESC LIMIT 1",
                                  (o,)).fetchone()
                    if l:
                        c.execute("DELETE FROM tampon WHERE id=?", (l["id"],))
                        return {"tip": "gerial", "ham": l["ham"], "ses": "uyari"}
                    x = c.execute("SELECT id,ham FROM okutma WHERE oturum=? ORDER BY id DESC LIMIT 1",
                                  (o,)).fetchone()
                    if x:
                        c.execute("DELETE FROM okutma WHERE id=?", (x["id"],))
                        return {"tip": "gerial", "ham": x["ham"], "ses": "uyari"}
                    return {"tip": "bos", "ses": "uyari"}
                if k == "fazla":
                    hs = [r["ham"] for r in c.execute(
                        "SELECT ham FROM tampon WHERE oturum=? ORDER BY id", (o,))]
                    c.execute("DELETE FROM tampon WHERE oturum=?", (o,))
                    for h in hs or [""]:
                        c.execute("INSERT INTO okutma(oturum,ts,ham,miktar,tip,not_) "
                                  "VALUES(?,?,?,1,'fazla','elle işaretlendi')", (o, ts, h))
                    return {"tip": "fazla_elle", "barkodlar": hs, "ses": "uyari"}
                if k == "atla":
                    hs = [r["ham"] for r in c.execute(
                        "SELECT ham FROM tampon WHERE oturum=? ORDER BY id", (o,))]
                    c.execute("DELETE FROM tampon WHERE oturum=?", (o,))
                    if hs:
                        c.execute("INSERT INTO kuyruk(oturum,ts,barkodlar,raf) VALUES(?,?,?,?)",
                                  (o, ts, json.dumps(hs, ensure_ascii=False), d.get("raf")))
                    return {"tip": "kuyruk", "barkodlar": hs, "ses": "kuyruk"}
                if k == "bitir":
                    grup_coz(c, o, amb, d.get("raf"))
                    c.execute("UPDATE oturum SET bitir=? WHERE id=?", (ts, o))
                    return {"tip": "bitti", "ses": "ok"}

            if U.startswith("##RAF-"):
                return {"tip": "raf", "raf": U[6:-2], "ses": "ok"}

            c.execute("INSERT INTO tampon(oturum,ts,ham) VALUES(?,?,?)", (o, ts, ham))
            n = c.execute("SELECT COUNT(*) n FROM tampon WHERE oturum=?", (o,)).fetchone()["n"]
            sayilan = set(r["beklenen_id"] for r in c.execute(
                "SELECT beklenen_id FROM okutma WHERE oturum=? AND beklenen_id IS NOT NULL", (o,)))
            r = coz(c, ham, amb, sayilan)
            return {"tip": "tampon", "adet": n, "ham": ham, "coz": r["t"],
                    "kod": r.get("kod"), "aciklama": r.get("aciklama"), "ses": "tik"}

        if p == "/api/kuyruk":
            rs = c.execute("SELECT * FROM kuyruk WHERE oturum=? AND cozuldu=0",
                           (d["oturum"],)).fetchall()
            return {"kuyruk": [{"id": r["id"], "barkodlar": json.loads(r["barkodlar"]),
                                "raf": r["raf"], "ts": r["ts"][:19]} for r in rs]}

        if p == "/api/kuyruk_coz":
            q = c.execute("SELECT * FROM kuyruk WHERE id=?", (d["id"],)).fetchone()
            hs = json.loads(q["barkodlar"])
            b = c.execute("SELECT * FROM beklenen WHERE id=?", (d["bid"],)).fetchone()
            for h in hs:
                c.execute("INSERT OR REPLACE INTO eslesme VALUES(?,?,?,?)",
                          (norm(h), b["kod"], "", ts))
            c.execute("""INSERT INTO okutma(oturum,ts,ham,kod,seri,miktar,beklenen_id,tip,not_)
                         VALUES(?,?,?,?,?,1,?,'eslesti','kuyruktan çözüldü')""",
                      (q["oturum"], ts, " + ".join(hs), b["kod"], b["seri"], b["id"]))
            c.execute("UPDATE kuyruk SET cozuldu=1 WHERE id=?", (d["id"],))
            return {"ok": 1}

        if p == "/api/bagla":
            o, bid, ham = d["oturum"], d["id"], d["ham"]
            b = c.execute("SELECT * FROM beklenen WHERE id=?", (bid,)).fetchone()
            c.execute("INSERT INTO okutma(oturum,ts,ham,kod,seri,miktar,beklenen_id,tip,not_)"
                      " VALUES(?,?,?,?,?,1,?,'eslesti','elle bağlandı')",
                      (o, ts, ham, b["kod"], b["seri"], bid))
            if d.get("ogren"):
                c.execute("INSERT OR REPLACE INTO eslesme VALUES(?,?,?,?)",
                          (norm(ham), b["kod"], b["seri"], ts))
            return {"ok": 1}

        if p == "/api/miktar":
            o = d["oturum"]
            amb = c.execute("SELECT ambar FROM oturum WHERE id=?", (o,)).fetchone()["ambar"]
            b = c.execute("SELECT * FROM beklenen WHERE kod=? AND ambar=?",
                          (d["kod"], amb)).fetchone()
            c.execute("INSERT INTO okutma(oturum,ts,ham,kod,seri,miktar,beklenen_id,tip,not_)"
                      " VALUES(?,?,?,?,?,?,?,'kod','miktar girildi')",
                      (o, ts, d["ham"], d["kod"], b["seri"] if b else "",
                       d["miktar"], b["id"] if b else None))
            return {"ok": 1}

        if p == "/api/fazla":
            c.execute("INSERT INTO okutma(oturum,ts,ham,miktar,tip) VALUES(?,?,?,1,'fazla')",
                      (d["oturum"], ts, d["ham"]))
            return {"ok": 1}

        if p == "/api/ara":
            q = f"%{d['q']}%"
            rs = c.execute("""SELECT id,kod,aciklama,seri,kirli FROM beklenen
                              WHERE ambar=? AND (kod LIKE ? OR aciklama LIKE ?)
                              ORDER BY kirli DESC LIMIT 25""",
                           (d["ambar"], q, q)).fetchall()
            return {"sonuc": [dict(r) for r in rs]}

        if p == "/api/rapor":
            c.execute("UPDATE oturum SET bitir=? WHERE id=?", (ts, d["oturum"]))
            return {"ok": 1}

        return {"hata": "bilinmeyen istek"}


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "yukle":
        yukle(sys.argv[2])
    elif len(sys.argv) > 1 and sys.argv[1] == "rapor":
        rapor(int(sys.argv[2]), sys.argv[3])
    else:
        if not os.path.exists(DB):
            sys.exit("Once rapor yukleyin:  python depo_sayim.py yukle <rapor.xlsx>")
        socketserver.TCPServer.allow_reuse_address = True
        with socketserver.TCPServer(("", PORT), H) as s:
            print(f"Sayim arayuzu hazir ->  http://localhost:{PORT}")
            print("Durdurmak icin Ctrl+C")
            s.serve_forever()
