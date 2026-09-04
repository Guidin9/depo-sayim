"""7 sekmeli sayım raporu (CLAUDE.md 5, 12).

depo_sayim.py:307-375 taşındı. Sekme adları ve sütun düzeni korundu, üstüne:
  * Fazla / Eşleşen / Tiger Düzeltme sekmelerine Raf sütunu
  * Eksik sekmesi sayım dışı kalemleri listelemez, sayıyı dipnot olarak yazar
  * Lot ve izlemesiz kalemlerde adet farkı hesaplanır (seri takipli kalemlerde
    prototipteki davranış aynen korunur: okutulmayan satır eksiktir)
  * Etiketler sekmesi: kendi bastığımız etiketlerin defteri (CLAUDE.md 12)

rapor_verisi() hem Excel yazıcısını hem arayüzdeki sekme önizlemesini besler ki
iki yerde iki ayrı gerçek olmasın.
"""
import os

from . import matching
from .norm import kirli_mi

SEKME = ("Eksik", "Fazla", "Eşleşen", "Yedek Parça", "Tiger Düzeltme",
         "Barkod Tablosu", "Etiketler")

BASLIKLAR = {
    "Eksik": ["Malzeme Kodu", "Açıklama", "Beklenen Seri/Lot", "İzleme", "Miktar",
              "Birim", "Not"],
    # Ürün Adı: Tiger'da kaydı olmayan ürünlerde açıklama JOIN'den gelemez,
    # kullanıcı elle yazar (CLAUDE.md 5, DEMO_FEEDBACK.md 3).
    "Fazla": ["Zaman", "Raf", "Okutulan", "Malzeme Kodu", "Açıklama", "Seri",
              "Miktar", "Not", "Ürün Adı"],
    # Okutulan Barkodlar: grubun tamamı (`okutma.ham`). Denetim izinin asıl
    # değeri bu — "hangi cihazda hangi barkodu okuttum" sorusunun tek cevabı.
    "Eşleşen": ["Zaman", "Raf", "Malzeme Kodu", "Açıklama", "Seri/Lot",
                "Okutulan Barkodlar", "Miktar", "Tip", "Not"],
    # Yedek parçalar Tiger'da kayıtlı değil; fazla da değiller, eksik de.
    # Sayacı bozmasınlar diye kendi sekmelerinde dururlar (saha bildirimi I4).
    "Yedek Parça": ["Zaman", "Raf", "Okutulan Barkodlar", "Ürün Adı", "Adet",
                    "Not"],
    "Tiger Düzeltme": ["Malzeme Kodu", "Açıklama", "MEVCUT (hatalı) Seri No",
                       "YENİ (gerçek) Seri No", "Raf", "Zaman"],
    "Barkod Tablosu": ["Okutulan Barkod", "Malzeme Kodu", "Açıklama",
                       "Öğrenildiği An"],
    "Etiketler": ["Etiket", "Tür", "Malzeme Kodu", "Açıklama",
                  "Bağlandığı Kayıt", "Raf", "Basıldığı An", "Bağlandığı An"],
}

DIPNOT = {
    "Yedek Parça": ["Yedek parça modunda okutulanlar. Tiger'da aranmadılar — "
                    "eksik ya da fazla sayılmazlar.",
                    "Tiger'a sayım fazlası fişi olarak GİRİLMEZLER; ne "
                    "yapılacağına ayrıca karar verin."],
    "Tiger Düzeltme": ["Bu sayfadaki kayıtlar Tiger'da seri numarası düzeltmesi "
                       "gerektirir.", "Ambar Sayımı ekranından fiş oluştururken "
                       "kullanın."],
    "Barkod Tablosu": ["Bu barkodları Tiger'da malzeme kartı > Birimler > Barkod "
                       "alanına yazın.", "Yazdıktan sonra bu ürünler sorusuz eşleşir.",
                       "Liste bu raporun ambarındaki malzemelerle sınırlıdır; "
                       "önceki oturumlarda öğrenilenler de burada — Tiger'a "
                       "girilmemiş olabilirler.",
                       "DİKKAT: bu listede CİHAZA ÖZEL seri numaraları da "
                       "bulunabilir. Kirli bir kaydı düzeltmek için okutulan "
                       "gerçek S/N de öğrenilir (kutudaki bütün barkodlar "
                       "kaydedilsin diye). Malzeme kartına yazılacak barkod "
                       "ÜRÜN TİPİNE ait olmalı — o malzemenin her adedinde "
                       "aynı olan barkod (UPC / üretici P/N). Tek bir cihazda "
                       "bulunan numarayı yazmayın; Tiger Düzeltme sekmesinde "
                       "zaten seri no olarak duruyor."],
    "Etiketler": ["Kendi bastığımız etiketlerin defteri: hangi numara neye yapıştı.",
                  "Tiger'a yazılacak değerler Tiger Düzeltme ve Barkod Tablosu "
                  "sekmelerinde; bu sayfa fiziksel etiketi bulmak içindir.",
                  "Malzemesi boş satırlar henüz kullanılmamış, havuzda bekleyen "
                  "etiketlerdir.",
                  "Kutu (DK-) satırlarında malzeme kabın İÇERİĞİDİR ve "
                  "değişebilir; etiket numarası değişmez. Adet burada "
                  "yazmaz — her sayımda yeniden sorulur."],
}


ETIKET_TUR_ADI = {"malzeme": "Malzeme", "seri": "Seri", "kutu": "Kutu"}


def _kisa(ts):
    return (ts or "")[:19].replace("T", " ")


def _yeni_seri(ham):
    """Tiger'a yazılacak gerçek seri numarasını seçer.

    Kuyruktan çözülen grupta okutma 'A + B' biçiminde saklanır (denetim izi).
    Tiger Düzeltme sekmesine tek bir değer yazılmalı: perakende barkodu olmayan
    en uzun parça — grup çözümlemesindeki kuralın aynısı (matching.grup_coz).

    Kendi bastığımız etiketler son sıraya düşer. 'DM-000123 + DS-000045'
    grubunda iki parça da aynı uzunlukta olduğu için max() ilkini — yani MALZEME
    etiketini — seçerdi; malzeme etiketi seri numarası değildir. Seri etiketi
    ancak başka aday yoksa kullanılır (grup_coz'daki yeni_sn sırasının aynısı).

    ELEMELER TEK PARÇALI GİRDİDE DE YAPILIR — `matching._sn_karar` ile aynı
    sözleşme, aday kalmazsa öneri YOKTUR (boş dize).
    """
    from .etiketler import etiket_turu
    from .norm import upc_mi
    parcalar = [p.strip() for p in str(ham or "").split(" + ") if p.strip()]
    # İki eleme, ikisi de KOŞULSUZ — kısa devrenin ÜSTÜNDE.
    #
    # KAP KODU: DK-000007 bir cihazın kimliği değil, durduğu kabın numarasıdır.
    # Tiger'a "bu cihazın S/N'i DK-000007" demek, kap ertesi ay boşaldığında
    # hiçbir şeye karşılık gelmeyen bir seri numarası bırakır.
    #
    # UPC: perakende barkodu o malzemenin HER ADEDİNDE aynıdır; seri numarası
    # tek cihaza aittir. Bu eleme bir dönem yalnızca çok parçalı girdide
    # yapılıyordu ve tek parçalı dal B2'yi geri getiriyordu (2026-09-04
    # denetimi, DENETIM_20260904.md K1): `_fazla_seri` malzeme kodunu eledikten
    # sonra elinde yalnızca UPC kalan bir grupta perakende barkodunu Tiger'a
    # seri numarası diye yazdırıyordu — `kuyruk_coz`, `fazla_bagla`,
    # `kuyruk_fazla` ve `##FAZLA##`, yani `_sn_karar`'ın korumadığı DÖRT yol.
    # Rapordaki son savunma da tutmuyor: `kirli_mi("198701689928", kod)` TEMİZ
    # döner.
    #
    # `or parcalar` gibi bir geri düşüş YOK: hepsi elenirse öneri üretilmez.
    # `_sn_karar` de aynısını yapıyor (`if not temiz: return yedek`).
    parcalar = [p for p in parcalar
                if etiket_turu(p) != "kutu" and not upc_mi(p)]
    if not parcalar:
        return ""
    if len(parcalar) == 1:
        return parcalar[0]
    for sinif in (None, "seri"):            # önce gerçek S/N, sonra seri etiketi
        havuz = [p for p in parcalar if etiket_turu(p) == sinif]
        if havuz:
            return max(havuz, key=len)
    return max(parcalar, key=len)


def eksik_kayitlar(c, oturum_id):
    """Sayımda karşılığı bulunamayan beklenen satırlar.

    Rapordaki Eksik sekmesi ve sayım sonu eşleştirme ekranı aynı listeyi
    kullanır — iki yerde iki ayrı gerçek olmasın (DEMO_FEEDBACK.md 6).

    (eksik, adet_fazlasi, haric_sayisi) döner; ilk ikisi sözlük listesidir.
    """
    o = c.execute("SELECT * FROM oturum WHERE id=?", (oturum_id,)).fetchone()
    if not o:
        raise ValueError("Oturum #%s bulunamadı" % oturum_id)
    yukleme, ambar = o["yukleme"], o["ambar"]

    sayilan = {r["beklenen_id"]: r["adet"] for r in c.execute(
        "SELECT beklenen_id, SUM(miktar) adet FROM okutma WHERE oturum=? "
        "AND beklenen_id IS NOT NULL GROUP BY beklenen_id", (oturum_id,))}

    eksik, adet_fazlasi = [], []
    haric_sayisi = 0
    for r in c.execute("SELECT * FROM beklenen WHERE yukleme=? AND ambar=? ORDER BY id",
                       (yukleme, ambar)):
        if r["haric"]:
            haric_sayisi += 1
            continue
        okunan = sayilan.get(r["id"], 0)
        ortak = {"id": r["id"], "kod": r["kod"], "aciklama": r["aciklama"],
                 "seri": r["seri"], "izleme": r["izleme"], "birim": r["birim"],
                 "kirli": r["kirli"]}
        # TEK KURAL, seri ve lot için aynı: beklenen - sayılan.
        #
        # Seri takipli satır eskiden ayrı bir daldan geçiyordu ("okutulduysa
        # bitti"). Gerçek veride yanlış: `izleme='seri'` olduğu hâlde miktarı
        # 2 ve 4 olan 32 satır var. Tek okutma o satırı kapatıyor, ikinci
        # cihaz "tekrar" deyip sayılmıyor ve BURAYA da girmiyordu — adet ne
        # sayaçta, ne eksikte, ne fazlada görünüyordu.
        beklenen = matching.beklenen_adet(r)
        kirli_not = "KIRLI KAYIT — " + r["kirli_sebep"] if r["kirli"] else ""
        fark = beklenen - okunan
        if fark > 0:
            # Hiç okutulmamış satırda not yalnızca kirlilik bilgisidir
            # (prototipteki metin korunuyor); yarım kalmışta adet de yazılır.
            not_ = kirli_not if not okunan else (
                ("adet farkı — sayılan %g / beklenen %g" % (okunan, beklenen))
                + (" | " + kirli_not if kirli_not else ""))
            eksik.append(dict(ortak, miktar=fark, not_=not_))
        elif fark < 0:
            adet_fazlasi.append(dict(ortak, miktar=-fark,
                                     not_="adet fazlası — sayılan %g / beklenen %g"
                                          % (okunan, beklenen)))
    return eksik, adet_fazlasi, haric_sayisi


def rapor_verisi(c, oturum_id):
    """Sekme adı -> {basliklar, satirlar, dipnot} sözlüğü."""
    o = c.execute("SELECT * FROM oturum WHERE id=?", (oturum_id,)).fetchone()
    if not o:
        raise ValueError("Oturum #%s bulunamadı" % oturum_id)
    yukleme, ambar = o["yukleme"], o["ambar"]

    ham_eksik, ham_adet_fazlasi, haric_sayisi = eksik_kayitlar(c, oturum_id)
    eksik = [[e["kod"], e["aciklama"], e["seri"], e["izleme"], e["miktar"],
              e["birim"], e["not_"]] for e in ham_eksik]
    adet_fazlasi = [["", "", e["kod"], e["kod"], e["aciklama"], e["seri"],
                     e["miktar"], e["not_"], ""] for e in ham_adet_fazlasi]

    # Açıklama boşsa elle yazılan ad devreye girer: kodu olmayan kayıtta
    # sütunun boş kalması raporu okunamaz yapıyordu.
    fazla = [[_kisa(r["ts"]), r["raf"] or "", r["ham"], r["kod"] or "?",
              r["aciklama"] or r["ad"] or "", r["seri"] or "", r["miktar"],
              r["not_"] or "", r["ad"] or ""]
             for r in c.execute("""SELECT o.*, b.aciklama FROM okutma o
                                   LEFT JOIN beklenen b ON b.kod=o.kod
                                        AND b.yukleme=? AND b.ambar=?
                                   WHERE o.oturum=? AND o.tip='fazla' 
                                   GROUP BY o.id ORDER BY o.id""",
                                (yukleme, ambar, oturum_id))]
    for satir in adet_fazlasi:
        satir[0] = _kisa(o["bitir"] or o["basla"])
        fazla.append(satir)

    eslesen = [[_kisa(r["ts"]), r["raf"] or "", r["kod"], r["aciklama"] or "",
                r["seri"], r["ham"] or "", r["miktar"], r["tip"], r["not_"] or ""]
               for r in c.execute("""SELECT o.*, b.aciklama FROM okutma o
                                     LEFT JOIN beklenen b ON b.id=o.beklenen_id
                                     WHERE o.oturum=? AND o.tip IN ('eslesti','kod')
                                     ORDER BY o.id""", (oturum_id,))]

    # Tiger'a önerilecek seri numarası `okutma.yeni_seri`'den okunur — `ham`
    # artık grubun BÜTÜN barkodlarını taşıyor ve içinde malzeme kodu da var.
    # `_yeni_seri(ham)` ona bakarsa kodu seri no sanıp Tiger'a yazdırır
    # (ACIL_PLAN 3'te kapatılan hata).
    #
    # `yeni_seri IS NULL` = bu sütun eklenmeden önce yazılmış eski kayıt; orada
    # `ham` hâlâ tek değerdi, eski kural uygulanır. Karar boşsa satır rapora
    # GİRMEZ: `sn_yok` sözleşmesi (seri numarası verilmedi, düzeltme üretilmez).
    yedek = [[_kisa(r["ts"]), r["raf"] or "", r["ham"] or "", r["ad"] or "",
              r["miktar"], r["not_"] or ""]
             for r in c.execute("""SELECT * FROM okutma WHERE oturum=? AND tip='yedek'
                                   ORDER BY id""", (oturum_id,))]

    duzeltme = []
    duzeltme_elenen = 0
    duzeltme_tahmin = 0
    for r in c.execute("""SELECT o.ham, o.ts, o.raf, o.yeni_seri, o.sn_adaylar,
                          b.kod, b.aciklama,
                          b.seri FROM okutma o
                          JOIN beklenen b ON b.id=o.beklenen_id
                          WHERE o.oturum=? AND b.kirli=1
                          ORDER BY o.id""", (oturum_id,)):
        yeni = (r["yeni_seri"] if r["yeni_seri"] is not None
                else _yeni_seri(r["ham"]))
        yeni = (yeni or "").strip()
        if not yeni:
            continue
        # SON SAVUNMA: bu sayfanın TEK işi Tiger'daki kirli seri numaralarını
        # temizlemek. Önerilen değerin kendisi kirliyse öneri değil yeni bir
        # kirlilik olur — özellikle "malzeme koduyla başlıyor" deseni
        # (`kirli_mi` -> kod+sayac).
        #
        # Bu ağ boşuna değil: 2026-08-27'de `eslesti` dalı `yeni_seri`'yi NULL
        # bırakınca rapor eski kurala düşüp `ham`'daki MALZEME KODUNU
        # öneriyordu. Motor düzeltildi; ağ, bir sonraki benzer hatanın Tiger'a
        # ulaşmasını engellemek için duruyor.
        if kirli_mi(yeni, r["kod"])[0]:
            duzeltme_elenen += 1
            continue
        # `sn_adaylar` hâlâ dolu = kullanıcı seçmedi, değer uygulamanın tahmini.
        if r["sn_adaylar"]:
            duzeltme_tahmin += 1
        duzeltme.append([r["kod"], r["aciklama"], r["seri"], yeni,
                         r["raf"] or "", _kisa(r["ts"])])

    # Yalnızca BU raporun ambarındaki malzemelere ait barkodlar.
    #
    # Eskiden `eslesme` tablosunun tamamı dökülüyordu: başka ambarların
    # malzemeleri, önceki sayımlardan öğrenilen her şey ve basılmış her `DM-`
    # etiketi her raporda yeniden çıkıyor, liste sonsuza kadar büyüyordu.
    # Sekmenin işi "Tiger'da HANGİ malzeme kartına ne yazacağım" — o kart bu
    # ambarın dışındaysa bu raporun işi değil.
    barkodlar = [[r["barkod"], r["kod"], r["aciklama"] or "", _kisa(r["ts"])]
                 for r in c.execute("""SELECT e.barkod, e.kod, e.ts,
                        (SELECT aciklama FROM beklenen WHERE kod=e.kod
                         AND yukleme=? AND ambar=? LIMIT 1) aciklama
                        FROM eslesme e
                        WHERE EXISTS(SELECT 1 FROM beklenen b WHERE b.kod=e.kod
                                     AND b.yukleme=? AND b.ambar=? AND b.haric=0)
                        ORDER BY e.kod, e.barkod""",
                                    (yukleme, ambar, yukleme, ambar))]

    # Üç tür de kendi adıyla çıkar. Eskiden "malzeme değilse Seri" yazıyordu ve
    # üçüncü sınıf eklendiği anda kap etiketleri raporda "Seri" görünürdü
    # (KUTU_TASARIM.md 4). Kap satırında malzeme `etiket` tablosunda değil
    # `kutu` defterinde durur — kabın içeriği değişebilir, etiket numarası
    # değişmez.
    etiket_satir = [[r["gosterim"], ETIKET_TUR_ADI.get(r["tur"], r["tur"] or "?"),
                     r["malzeme"] or r["kutu_malzeme"] or "", r["aciklama"] or "",
                     r["slot"] or "", r["raf"] or "", _kisa(r["ts"]),
                     _kisa(r["ts_bagla"] or r["kutu_guncelle"])]
                    for r in c.execute("""SELECT e.*,
                          (SELECT malzeme FROM kutu WHERE kod=e.kod) kutu_malzeme,
                          (SELECT ts_guncelle FROM kutu WHERE kod=e.kod) kutu_guncelle,
                          (SELECT aciklama FROM beklenen
                           WHERE kod=COALESCE(e.malzeme,
                                 (SELECT malzeme FROM kutu WHERE kod=e.kod))
                           ORDER BY yukleme DESC LIMIT 1) aciklama,
                          (SELECT seri FROM beklenen WHERE id=e.beklenen_id) slot
                          FROM etiket e ORDER BY e.tur, e.kod""")]

    veri = {}
    for ad, satirlar in (("Eksik", eksik), ("Fazla", fazla), ("Eşleşen", eslesen),
                         ("Yedek Parça", yedek),
                         ("Tiger Düzeltme", duzeltme), ("Barkod Tablosu", barkodlar),
                         ("Etiketler", etiket_satir)):
        veri[ad] = {"basliklar": BASLIKLAR[ad], "satirlar": satirlar,
                    "dipnot": list(DIPNOT.get(ad, []))}
    if duzeltme_elenen:
        veri["Tiger Düzeltme"]["dipnot"].append(
            "%d öneri elendi: önerilen seri numarasının kendisi kirli desene "
            "uyuyordu (malzeme koduyla başlıyor, boşluk içeriyor vb.). Bu "
            "kayıtlar Tiger'da düzeltilmeden kaldı — sayımları işlendi."
            % duzeltme_elenen)
    if duzeltme_tahmin:
        veri["Tiger Düzeltme"]["dipnot"].append(
            "%d satırda seri numarası UYGULAMA TARAFINDAN SEÇİLDİ: üründe birden "
            "çok tanınmayan barkod vardı ve hangisinin cihaza özel olduğu "
            "sorulmadan en uzunu alındı. Tiger'a girmeden önce bu satırlara "
            "bakın." % duzeltme_tahmin)
    if haric_sayisi:
        veri["Eksik"]["dipnot"].append(
            "%d kalem sayım dışı filtresiyle çıkarıldı (lisans / hizmet / nakliye "
            "vb.) — eksik sayılmadılar." % haric_sayisi)
    veri["_ozet"] = {
        "oturum": oturum_id, "ambar": ambar, "yukleme": yukleme,
        "basla": _kisa(o["basla"]), "bitir": _kisa(o["bitir"]), "durum": o["durum"],
        "haric": haric_sayisi,
        "sayilar": {ad: len(veri[ad]["satirlar"]) for ad in SEKME},
    }
    return veri


def excel_yaz(c, oturum_id, yol):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    veri = rapor_verisi(c, oturum_id)
    wb = Workbook()
    wb.remove(wb.active)
    kalin = Font(name="Arial", bold=True, color="FFFFFF")
    dolgu = PatternFill("solid", fgColor="2A3140")
    not_yazi = Font(name="Arial", italic=True, color="666666")

    for ad in SEKME:
        s = veri[ad]
        ws = wb.create_sheet(ad)
        ws.append(s["basliklar"])
        for h in ws[1]:
            h.font = kalin
            h.fill = dolgu
            h.alignment = Alignment(vertical="center")
        for satir in s["satirlar"]:
            ws.append(list(satir))
        for i, b in enumerate(s["basliklar"], 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = \
                max(12, min(45, len(b) + 18))
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        if s["dipnot"]:
            ws.append([])
            for d in s["dipnot"]:
                ws.append([d])
                ws.cell(ws.max_row, 1).font = not_yazi

    if os.path.dirname(yol):
        os.makedirs(os.path.dirname(os.path.abspath(yol)), exist_ok=True)
    wb.save(yol)
    return veri["_ozet"]


def rapor_yolu(oturum_id, klasor=None):
    from .db import VERI
    klasor = klasor or os.path.join(VERI, "rapor")
    return os.path.join(klasor, "sayim_%s.xlsx" % oturum_id)
