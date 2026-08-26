"""Tiger raporu yükleyici (Excel + JSON) ve sayım dışı kalem filtresi.

Kaynaklar (CLAUDE.md 3.1 / 3.2):
  * Lot / Seri Envanter Raporu  -> kaynak='seri_lot'  (izleme: seri | lot)
  * Envanter Raporu             -> kaynak='envanter'  (izleme: yok, adet bazlı)

Excel'de başlık satırı 2. satırdadır (1. satır rapor başlığı), bu yüzden başlık
satırı 'Malzeme Kodu' geçen ilk satır olarak aranır — prototipteki davranış
(depo_sayim.py:91-97).
"""
import datetime
import json
import os

from .norm import izleme_coz, kirli_mi, norm, sifirsiz

# Normalize edilmiş başlık -> alan adı. norm() Türkçe karakteri katladığı için
# 'İzleme Yöntemi', 'IZLEME YONTEMI' ve 'izleme_yontemi' aynı anahtara düşer.
#
# Sütun sırası, sütun sayısı ve başlık satırının kaçıncı satırda olduğu önemli
# değil — eşleştirme başlık ADINA göre yapılır. Tanınmayan sütunlar yoksayılır,
# bu yüzden farklı ambarların/dönemlerin raporları da okunur.
BASLIK = {
    "MALZEMETURU": "tur",
    "MALZEMEKODU": "kod",
    "STOKKODU": "kod",
    "MALZEMEACIKLAMASI": "aciklama",
    "MALZEMEADI": "aciklama",
    "STOKADI": "aciklama",
    "IZLEMEYONTEMI": "izleme_ham",
    "AMBARMALIYETGRUBU": "ambar",
    "AMBARNO": "ambar_no",
    "AMBARKODU": "ambar_no",
    "SERILOTNO": "seri",
    "SERINO": "seri",
    "LOTNO": "seri",
    "SERILOTACIKLAMASI": "seri_aciklama",
    "ENVANTERMIKTARI": "miktar",
    "MIKTAR": "miktar",
    "BIRIM": "birim",
    "ANABIRIM": "birim",
}
# Ambar bilgisi iki sütundan birinden gelebilir; hangisi varsa o kullanılır.
ZORUNLU = ("kod",)
ZORUNLU_GRUP = (("ambar", "ambar_no"),)


class YuklemeHatasi(Exception):
    pass


def _sayi(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _metin(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


# ---------------------------------------------------------------- okuyucular
def _baslik_satiri(satir):
    """Satır başlık satırı mı? Malzeme kodu sütununu taşıyorsa evet."""
    if not satir:
        return None
    eslesme = [BASLIK.get(norm(x)) for x in satir]
    return eslesme if "kod" in eslesme else None


def excel_satirlar(yol):
    """(alan kümesi, satır sözlükleri) döner.

    Başlık satırı aranarak bulunur (Tiger'da 2. satırdadır ama sabitlenmedi) ve
    sayfalar sırayla denenir — rapor ilk sayfada olmayabilir.
    """
    from openpyxl import load_workbook
    wb = load_workbook(yol, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            it = ws.iter_rows(values_only=True)
            basliklar = None
            for r in it:
                basliklar = _baslik_satiri(r)
                if basliklar:
                    break
            if not basliklar:
                continue
            alanlar = set(a for a in basliklar if a)
            satirlar = []
            for r in it:
                d = {}
                for i, alan in enumerate(basliklar):
                    if alan and i < len(r):
                        d[alan] = r[i]
                if _metin(d.get("kod")):
                    satirlar.append(d)
            if satirlar:
                return alanlar, satirlar
    finally:
        wb.close()
    raise YuklemeHatasi("Malzeme kodu sütunu bulunamadı — doğru rapor mu?")


def json_satirlar(yol):
    """Tiger'ın JSON rapor çıktısı. Anahtarlar norm() üzerinden eşleştirilir."""
    with open(yol, encoding="utf-8-sig") as f:
        veri = json.load(f)
    if isinstance(veri, dict):
        liste = next((v for v in veri.values() if isinstance(v, list) and v
                      and isinstance(v[0], dict)), None)
        if liste is None:
            raise YuklemeHatasi("JSON içinde satır listesi bulunamadı.")
        veri = liste
    if not isinstance(veri, list) or not veri or not isinstance(veri[0], dict):
        raise YuklemeHatasi("JSON beklenen biçimde değil (satır listesi lazım).")
    alanlar, satirlar = set(), []
    for h in veri:
        d = {}
        for k, v in h.items():
            alan = BASLIK.get(norm(k))
            if alan:
                d[alan] = v
                alanlar.add(alan)
        if _metin(d.get("kod")):
            satirlar.append(d)
    return alanlar, satirlar


def oku(yol):
    if str(yol).lower().endswith(".json"):
        return json_satirlar(yol)
    return excel_satirlar(yol)


# ---------------------------------------------------------------- yükleme
def yukle(c, yol, yukleme_id=None, dosya_adi=None):
    """Dosyayı beklenen tablosuna yazar. Özet sözlüğü döner."""
    alanlar, satirlar = oku(yol)
    eksik = [a for a in ZORUNLU if a not in alanlar]
    eksik += ["/".join(g) for g in ZORUNLU_GRUP if not any(a in alanlar for a in g)]
    if eksik:
        raise YuklemeHatasi("Raporda şu sütunlar yok: " + ", ".join(eksik))
    if not satirlar:
        raise YuklemeHatasi("Raporda veri satırı yok.")

    kaynak = "seri_lot" if "seri" in alanlar else "envanter"
    ts = datetime.datetime.now().isoformat()
    ad = dosya_adi or os.path.basename(str(yol))

    if yukleme_id is None:
        yukleme_id = c.execute(
            "INSERT INTO yukleme(ts,dosya_adi,kaynak,satir) VALUES(?,?,?,0)",
            (ts, ad, kaynak)).lastrowid
    else:
        onceki = c.execute("SELECT * FROM yukleme WHERE id=?", (yukleme_id,)).fetchone()
        if not onceki:
            raise YuklemeHatasi("Yükleme #%s bulunamadı." % yukleme_id)
        c.execute("UPDATE yukleme SET ts=?, dosya_adi=?, kaynak=? WHERE id=?",
                  (ts, onceki["dosya_adi"] + " + " + ad,
                   "karma" if onceki["kaynak"] != kaynak else kaynak, yukleme_id))

    # Envanter raporu, seri/lot raporunda zaten olan malzemeyi tekrar getirir;
    # seri/lot kaydı kazanır (CLAUDE.md 2.4 — takipsiz kalemler orada gelmez).
    mevcut = set()
    if kaynak == "envanter":
        mevcut = {(r["kod"], r["ambar"]) for r in c.execute(
            "SELECT DISTINCT kod, ambar FROM beklenen WHERE yukleme=? AND izleme<>'yok'",
            (yukleme_id,))}

    n = atlanan = 0
    for d in satirlar:
        kod = _metin(d.get("kod"))
        # Ambar boşsa satır "?" ambarında toplanır — sessizce 1'e karışmasın,
        # Kurulum ekranında görünüp fark edilsin.
        ambar = _metin(d.get("ambar")) or _metin(d.get("ambar_no")) or "?"
        if (kod, ambar) in mevcut:
            atlanan += 1
            continue
        seri = _metin(d.get("seri"))
        izleme = izleme_coz(d.get("izleme_ham")) if kaynak == "seri_lot" else "yok"
        if kaynak == "seri_lot" and izleme == "yok" and seri:
            izleme = "seri"  # izleme sütunu boş ama seri dolu
        kirli, sebep = kirli_mi(seri, kod) if izleme == "seri" else (0, "")
        sn = norm(seri)
        c.execute("""INSERT INTO beklenen(yukleme,kod,kod_n,aciklama,tur,ambar,izleme,
                     seri,seri_n,seri_n0,seri_aciklama,miktar,birim,kirli,kirli_sebep,
                     haric,haric_sebep,kaynak)
                     VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,0,'',?)""",
                  (yukleme_id, kod, norm(kod), _metin(d.get("aciklama")),
                   _metin(d.get("tur")), ambar, izleme, seri, sn, sifirsiz(sn),
                   _metin(d.get("seri_aciklama")), _sayi(d.get("miktar")),
                   _metin(d.get("birim")) or "AD", kirli, sebep, kaynak))
        n += 1

    c.execute("UPDATE yukleme SET satir=(SELECT COUNT(*) FROM beklenen WHERE yukleme=?) "
              "WHERE id=?", (yukleme_id, yukleme_id))
    haric_uygula(c, yukleme_id)
    c.commit()
    ozet = ozetle(c, yukleme_id)
    ozet["eklenen"] = n
    ozet["atlanan"] = atlanan
    ozet["kaynak"] = kaynak
    return ozet


# ---------------------------------------------------------------- sayım dışı
def _kural_tutar(kural, satir):
    d = norm(kural["desen"])
    if not d:
        return False
    if kural["tip"] == "tur":
        return d in norm(satir["tur"])
    return d in norm(satir["aciklama"])


def haric_uygula(c, yukleme_id):
    """Aktif kuralları beklenen.haric alanına yazar. Kural başına sayım döner."""
    kurallar = c.execute("SELECT * FROM haric_kural WHERE aktif=1 ORDER BY id").fetchall()
    sayim = {k["id"]: {"satir": 0, "adet": 0.0} for k in kurallar}
    c.execute("UPDATE beklenen SET haric=0, haric_sebep='' WHERE yukleme=?", (yukleme_id,))
    for r in c.execute("SELECT id,tur,aciklama,miktar FROM beklenen WHERE yukleme=?",
                       (yukleme_id,)).fetchall():
        for k in kurallar:
            if _kural_tutar(k, r):
                c.execute("UPDATE beklenen SET haric=1, haric_sebep=? WHERE id=?",
                          (k["tip"] + ":" + k["desen"], r["id"]))
                sayim[k["id"]]["satir"] += 1
                sayim[k["id"]]["adet"] += r["miktar"] or 0
                break
    c.commit()
    return sayim


def kural_ozeti(c, yukleme_id):
    """Kurulum ekranı için: her kural kaç satır / kaç adet etkiliyor.

    Kapalı kuralların etkisi de hesaplanır ki kullanıcı açmadan önce görsün.
    """
    kurallar = c.execute("SELECT * FROM haric_kural ORDER BY tip, id").fetchall()
    satirlar = c.execute("SELECT tur,aciklama,miktar FROM beklenen WHERE yukleme=?",
                         (yukleme_id,)).fetchall()
    out = []
    for k in kurallar:
        tutan = [r for r in satirlar if _kural_tutar(k, r)]
        out.append({"id": k["id"], "tip": k["tip"], "desen": k["desen"],
                    "aktif": bool(k["aktif"]), "varsayilan": bool(k["varsayilan"]),
                    "satir": len(tutan),
                    "adet": sum(r["miktar"] or 0 for r in tutan)})
    return out


# ---------------------------------------------------------------- özet
def ozetle(c, yukleme_id):
    y = c.execute("SELECT * FROM yukleme WHERE id=?", (yukleme_id,)).fetchone()
    izl = c.execute("""SELECT izleme, COUNT(*) satir, SUM(miktar) adet,
                       COUNT(DISTINCT kod) malzeme, SUM(kirli) kirli
                       FROM beklenen WHERE yukleme=? GROUP BY izleme""",
                    (yukleme_id,)).fetchall()
    amb = c.execute("""SELECT ambar, COUNT(*) satir, SUM(miktar) adet
                       FROM beklenen WHERE yukleme=? GROUP BY ambar ORDER BY 2 DESC""",
                    (yukleme_id,)).fetchall()
    sebep = c.execute("""SELECT kirli_sebep sebep, COUNT(*) satir FROM beklenen
                         WHERE yukleme=? AND kirli=1 GROUP BY kirli_sebep
                         ORDER BY 2 DESC""", (yukleme_id,)).fetchall()
    har = c.execute("SELECT COUNT(*) satir, COALESCE(SUM(miktar),0) adet FROM beklenen "
                    "WHERE yukleme=? AND haric=1", (yukleme_id,)).fetchone()
    # Raporda GERÇEKTEN geçen malzeme türleri. Sayım dışı kurallarının yarısı
    # `tur` üzerinden çalışıyor ama örnek Ambar 1 çıktısında bu sütun `TM`/`TK`
    # kısa kodlarını döndürüyor — varsayılan desenlerin (`YAZILIM`, `HİZMET`…)
    # hiçbiri tutmuyor ve kullanıcı NEDEN tutmadığını göremiyordu.
    tur = c.execute("""SELECT COALESCE(NULLIF(TRIM(tur),''),'(boş)') tur,
                       COUNT(*) satir FROM beklenen WHERE yukleme=?
                       GROUP BY 1 ORDER BY 2 DESC LIMIT 20""",
                    (yukleme_id,)).fetchall()
    return {
        "etiket_cakisma": etiket_cakismasi(c, yukleme_id),
        "yukleme": yukleme_id,
        "dosya": y["dosya_adi"],
        "ts": y["ts"],
        "kaynak": y["kaynak"],
        "satir": y["satir"],
        "izleme": [dict(r) for r in izl],
        "ambarlar": [dict(r) for r in amb],
        "kirli_sebep": [dict(r) for r in sebep],
        "kirli": sum(r["kirli"] or 0 for r in izl),
        "haric": {"satir": har["satir"], "adet": har["adet"]},
        "turler": [dict(r) for r in tur],
    }


def etiket_cakismasi(c, yukleme_id):
    """Tiger verisinde bizim etiket desenimize benzeyip deftere ait olmayan kayıt.

    Kendi bastığımız numaralar (DM-000123 / DS-000045) Tiger'a yazıldıktan
    sonraki yıl normal kod/seri olarak geri gelir — beklenen ve istenen budur.
    Ama defterde karşılığı OLMAYAN böyle bir değer varsa iki ihtimal var:
    ya defter kayboldu (data/etiket silinmiş), ya da tesadüfen aynı desende bir
    üretici kodu var. İkisi de sessiz kalırsa aynı numara ikinci kez basılabilir.
    """
    from . import etiketler
    cak = []
    for r in c.execute("SELECT kod, kod_n, seri, seri_n FROM beklenen WHERE yukleme=?",
                       (yukleme_id,)):
        for alan, ham, n in (("kod", r["kod"], r["kod_n"]),
                             ("seri", r["seri"], r["seri_n"])):
            if not n or not etiketler.etiket_mi(n):
                continue
            if c.execute("SELECT 1 FROM etiket WHERE kod=?", (n,)).fetchone():
                continue
            cak.append({"alan": alan, "deger": ham, "kod_n": n})
    return cak[:50]
