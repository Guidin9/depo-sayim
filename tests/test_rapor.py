"""5 sekmeli rapor (CLAUDE.md 5)."""
from openpyxl import load_workbook

from app import matching, reports
from tests.conftest import haric_kur, oturum_taze

SONRAKI = "##SONRAKI##"


def _satirlar(veri, sekme):
    return veri[sekme]["satirlar"]


def _bul(satirlar, deger, sutun=0):
    return [s for s in satirlar if s[sutun] == deger]


def test_sekme_adlari_ve_basliklar(c, ot):
    veri = reports.rapor_verisi(c, ot["id"])
    assert [s for s in veri if not s.startswith("_")] == list(reports.SEKME)
    assert veri["Fazla"]["basliklar"][1] == "Raf"
    assert veri["Eşleşen"]["basliklar"][1] == "Raf"
    assert veri["Tiger Düzeltme"]["basliklar"][4] == "Raf"


def test_eslesen_ve_eksik(c, ot, yaz):
    yaz("##RAF-A1##", "5S47WC2", SONRAKI)
    veri = reports.rapor_verisi(c, ot["id"])

    es = _bul(_satirlar(veri, "Eşleşen"), "210-ACXU-TİP2", 2)
    assert len(es) == 1 and es[0][1] == "A1" and es[0][4] == "5S47WC2"
    # okutulan seri artık Eksik'te değil, aynı malzemenin diğer serileri orada
    eksik = _satirlar(veri, "Eksik")
    assert not _bul(eksik, "5S47WC2", 2)
    assert _bul(eksik, "5SBGWC2", 2)


def test_tiger_duzeltme_sekmesi(c, ot, yaz):
    """Kirli slot doldurulunca eski/yeni seri no eşleşmesi rapora düşer."""
    yaz("##RAF-C2##", "0WGP72", "W3S2000G7745", SONRAKI)
    d = _satirlar(reports.rapor_verisi(c, ot["id"]), "Tiger Düzeltme")
    assert len(d) == 1
    assert d[0][0] == "0WGP72"
    assert d[0][2] == "0WGP72SAYIM1"      # MEVCUT (hatalı)
    assert d[0][3] == "W3S2000G7745"      # YENİ (gerçek)
    assert d[0][4] == "C2"


def test_barkod_tablosu(c, ot, yaz):
    yaz("0WGP72", "W3S2000G7745", SONRAKI)
    b = _satirlar(reports.rapor_verisi(c, ot["id"]), "Barkod Tablosu")
    assert ["W3S2000G7745", "0WGP72"] == b[0][:2]
    assert "SSD" in b[0][2]


def test_fazla_sekmesi(c, ot, yaz):
    """Fazla, ancak kullanıcı onayladıktan sonra rapora düşer."""
    r = yaz("##RAF-B1##", "210-BEJO", "YENISERI12345", SONRAKI)
    assert r["tip"] == "onay"
    assert _satirlar(reports.rapor_verisi(c, ot["id"]), "Fazla") == []

    matching.kuyruk_fazla(c, r["kuyruk_id"])
    f = _satirlar(reports.rapor_verisi(c, ot["id"]), "Fazla")
    assert len(f) == 1 and f[0][1] == "B1" and f[0][3] == "210-BEJO"


def test_elle_fazlada_urun_adi(c, ot, yaz):
    """Tiger'da kaydı olmayan ürün: kod yok, adı kullanıcı yazar.

    DEMO_FEEDBACK.md 3 — isimsiz fazla kaydı sonradan işe yaramıyordu.
    """
    r = yaz("KAYITSIZ-URUN-9911", "##FAZLA##")
    assert r["tip"] == "fazla_elle" and len(r["okutma"]) == 1
    c.execute("UPDATE okutma SET ad=? WHERE id=?", ("Kırmızı HP güç kablosu",
                                                    r["okutma"][0]))

    veri = reports.rapor_verisi(c, ot["id"])
    f = _satirlar(veri, "Fazla")
    assert veri["Fazla"]["basliklar"][8] == "Ürün Adı"
    assert len(f) == 1
    assert f[0][3] == "?"                          # malzeme kodu yok
    assert f[0][4] == "Kırmızı HP güç kablosu"     # açıklama adla doluyor
    assert f[0][8] == "Kırmızı HP güç kablosu"


def test_lot_adet_farki(c, ot, yaz):
    """Lot kaleminde 77 beklenirken 2 okutulduysa Eksik'te 75 adet fark görünür."""
    yaz("0C5RNH", SONRAKI)
    yaz("0C5RNH", SONRAKI)
    eksik = _bul(_satirlar(reports.rapor_verisi(c, ot["id"]), "Eksik"), "0C5RNH")
    assert len(eksik) == 1
    assert eksik[0][4] == 75
    assert "sayılan 2 / beklenen 77" in eksik[0][6]


def test_lot_adet_fazlasi(c, ot, yaz):
    """Beklenenden fazla okutulan lot kalemi Fazla sekmesine düşer."""
    b = c.execute("SELECT * FROM beklenen WHERE izleme='lot' AND miktar=1 "
                  "LIMIT 1").fetchone()
    for _ in range(3):
        yaz(b["kod"], SONRAKI)
    f = _bul(_satirlar(reports.rapor_verisi(c, ot["id"]), "Fazla"), b["kod"], 3)
    assert len(f) == 1 and f[0][6] == 2
    assert "adet fazlası" in f[0][7]


def test_haric_kalem_eksikte_yok_dipnotta_var(c, ot):
    """Sayım dışı kalem Eksik sekmesine girmez, sayısı dipnotta durur.

    Kural veriye gerçekten uyan bir desenle kuruluyor: varsayılanlar bu veride
    hiçbir satır yakalamıyor ve yakalamaMAlı da (bkz. tests/test_haric.py).
    """
    _, satir, kod = haric_kur(c)
    veri = reports.rapor_verisi(c, ot["id"])
    assert not _bul(_satirlar(veri, "Eksik"), kod)
    assert any("sayım dışı" in d for d in veri["Eksik"]["dipnot"])
    assert veri["_ozet"]["haric"] == satir


def test_excel_dosyasi_yazilir(c, ot, yaz, tmp_path):
    yaz("##RAF-A1##", "5S47WC2", SONRAKI)
    yaz("0WGP72", "W3S2000G7745", SONRAKI)
    yol = str(tmp_path / "rapor.xlsx")
    ozet = reports.excel_yaz(c, ot["id"], yol)

    wb = load_workbook(yol)
    assert wb.sheetnames == list(reports.SEKME)
    ws = wb["Tiger Düzeltme"]
    assert [x.value for x in ws[1]] == reports.BASLIKLAR["Tiger Düzeltme"]
    assert ws.cell(2, 3).value == "0WGP72SAYIM1"
    assert ws.freeze_panes == "A2"
    # dipnot satırları en altta duruyor
    assert any("Ambar Sayımı" in str(r[0].value) for r in ws.iter_rows(min_col=1,
                                                                      max_col=1))
    assert ozet["sayilar"]["Eşleşen"] == 2
    wb.close()


def test_gecmis_oturum_raporu_yeni_yuklemeden_etkilenmez(c, ot, yaz, tmp_path):
    """Yeni rapor yüklense de eski oturumun raporu aynı kalmalı."""
    yaz("5S47WC2", SONRAKI)
    onceki = reports.rapor_verisi(c, ot["id"])["_ozet"]["sayilar"]

    from app import importer
    import json
    yol = tmp_path / "yeni.json"
    yol.write_text(json.dumps([{"Malzeme Kodu": "YENI-1", "Ambar Maliyet Grubu": "1",
                                "Envanter Miktarı": 5}], ensure_ascii=False),
                   encoding="utf-8")
    importer.yukle(c, str(yol))          # yeni yükleme (#2)

    assert reports.rapor_verisi(c, ot["id"])["_ozet"]["sayilar"] == onceki


def test_kuyruktan_cozulen_tek_seri_yazar(c, ot, yaz):
    """Kuyruk 'A + B' olarak saklanır ama Tiger'a tek gerçek S/N yazılmalı."""
    yaz("198701689928", "EDBP0153231475674", "##SONRAKI##")
    kid = c.execute("SELECT id FROM kuyruk WHERE oturum=?", (ot["id"],)).fetchone()["id"]
    hedef = c.execute("SELECT id FROM beklenen WHERE seri='0WGP72SAYIM1'").fetchone()["id"]
    matching.kuyruk_coz(c, kid, hedef)

    d = _satirlar(reports.rapor_verisi(c, ot["id"]), "Tiger Düzeltme")
    assert d[0][2] == "0WGP72SAYIM1"
    assert d[0][3] == "EDBP0153231475674"      # UPC değil, en uzun aday
    # denetim izi Eşleşen sekmesinde tam haliyle duruyor
    es = _satirlar(reports.rapor_verisi(c, ot["id"]), "Eşleşen")
    assert "198701689928 + EDBP0153231475674" in [s[4] for s in es] or any(
        "kuyruktan" in str(s[7]) for s in es)


def test_arama_sayilani_isaretler(c, ot, yaz):
    yaz("0WGP72", "W3S2000G7745", "##SONRAKI##")     # SAYIM1 slotu doldu
    sonuc = matching.ara(c, ot["yukleme"], ot["ambar"], "0WGP72",
                         oturum=ot["id"])["satirlar"]
    sayim1 = next(s for s in sonuc if s["seri"] == "0WGP72SAYIM1")
    sayim2 = next(s for s in sonuc if s["seri"] == "0WGP72SAYIM2")
    assert sayim1["sayildi"] == 1 and sayim2["sayildi"] == 0
    assert sonuc.index(sayim2) < sonuc.index(sayim1)   # sayılmayanlar üstte
